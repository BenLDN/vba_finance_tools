#GUI implementation of the Reverse Pivot algorithm
#dfgfgdgf

from openpyxl import *
from tkinter import *
from tkinter.filedialog import askopenfilename, askdirectory

#Input file dialog box
def changeInput():
    inFile = askopenfilename(filetypes=(
        ("MS Excel Files", "*.xlsx"),
        ("All files", "*.*") ))
    inList.delete(END, 0)
    inList.insert(END, inFile)

#Output file has a standard name, but the directory has to be selected
def changeOutput():
    outDir = askdirectory()
    outFile = outDir + "/Output.xlsx"
    outList.delete(END, 0)
    outList.insert(END, outFile)

#Main Reverse Pivot function, the GUI was built around this
def convertFile():

    #getting data from the input boxes in the main window
    col_attrs=int(attrCol.get())
    row_attrs=int(attrRow.get())

    inFile=inList.get(0)
    outFile=outList.get(0)

    #load WB and select the active worksheet
    wb=load_workbook(inFile)
    ws=wb.active

    #mat is an array to store lines of the new, reverese pivoted data set
    #this is stored in memory before creating the new table
    
    mat=[]
    r2=0
    c2=0

    last_row=ws.max_row
    last_col=ws.max_column

    for c in range(col_attrs+1,last_col+1):
        for r in range(row_attrs+1,last_row+1):

            #app_line is a line in the new, reverse pivoted data set:
            #attributes follwoed by one single number
            
            app_line=[]

            for col_attr in range(1,col_attrs+1):
                app_line.append(ws.cell(row=r, column=col_attr).value)
            for row_attr in range(1,row_attrs+1):
                app_line.append(ws.cell(row=row_attr, column=c).value)
            
            app_line.append(ws.cell(row=r, column=c).value)
            
            mat.append(app_line)

    #creatinga  new workbook & worksheet, and then adding the data lines from mat[]
    #in the newly created sheet

    wb_new = Workbook()
    ws_new=wb_new.active
      
    for row in mat:
        r2+=1
        c2=0
        for cell in row:
            c2+=1
            ws_new.cell(row=r2, column=c2).value=cell

    wb_new.save(outFile)

#GUI

mGui=Tk()
ment=StringVar()

#Main window
mGui.geometry("780x300")
mGui.title("Reverse Pivot")

#input file and output file undefined
inFile=""
outFile=""

#listboxes show the current values of the input and output file
#reason for using a listbox:
#if more functionality is added, multiple input/output files might be needed
inputLabel=Label(mGui, text="Input File").grid(column=0, row=0, columnspan=2)
outputLabel=Label(mGui, text="Output File").grid(column=2, row=0, columnspan=2)

inList = Listbox(mGui, width=48, height=2)
inList.grid(column=0, row=1, columnspan=2)

outList = Listbox(mGui, width=48, height=2)
outList.grid(column=2, row=1, columnspan=2)

#Buttons call the change input/output file functions

changeInButton=Button(mGui, text="Change Input File...", command=changeInput).grid(column=0, row=2, columnspan=2, pady=20)
changeOutButton=Button(mGui, text="Change Output Directory...", command=changeOutput).grid(column=2, row=2, columnspan=2, pady=20)

#entry boxes for the number of column and row attributes
attrColLabel = Label(mGui, text="Number of column attributes: ").grid(column=1, row=4, sticky=E)
attrCol = Entry(mGui, width=2)
attrCol.grid(column=2, row=4, sticky=W, padx=40)

attrRowLabel = Label(mGui, text="Number of row attributes: ").grid(column=1, row=5, sticky=E)
attrRow = Entry(mGui, width=2)
attrRow.grid(column=2, row=5, sticky=W, padx=40)

#calling the main Reverse Pivot function
convertButton=Button(mGui, text="CONVERT", command=convertFile).grid(column=1, row=6, columnspan=2, pady=20)














































